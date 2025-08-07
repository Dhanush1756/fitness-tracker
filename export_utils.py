from io import BytesIO
from xhtml2pdf import pisa
from openpyxl import Workbook
from datetime import datetime

def generate_pdf_report(user, meals, workouts, weights):
    # Create HTML content
    html = f"""
    <html>
    <head>
        <title>Fitness Report for {user.name}</title>
        <style>
            body {{ font-family: Arial, sans-serif; }}
            h1 {{ color: #2c3e50; }}
            table {{ width: 100%; border-collapse: collapse; }}
            th, td {{ border: 1px solid #ddd; padding: 8px; text-align: left; }}
            th {{ background-color: #f2f2f2; }}
            .section {{ margin-bottom: 20px; }}
        </style>
    </head>
    <body>
        <h1>Fitness Report for {user.name}</h1>
        <p>Generated on {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>
        
        <div class="section">
            <h2>User Profile</h2>
            <p><strong>Name:</strong> {user.name}</p>
            <p><strong>Age:</strong> {user.age}</p>
            <p><strong>Gender:</strong> {user.gender}</p>
            <p><strong>Height:</strong> {user.height} cm</p>
            <p><strong>Weight:</strong> {user.weight} kg</p>
            <p><strong>Goal Weight:</strong> {user.goal_weight} kg</p>
            <p><strong>Daily Calorie Target:</strong> {user.daily_calories}</p>
        </div>
        
        <div class="section">
            <h2>Meal Logs</h2>
            <table>
                <tr>
                    <th>Date</th>
                    <th>Meal</th>
                    <th>Calories</th>
                    <th>Protein (g)</th>
                    <th>Carbs (g)</th>
                    <th>Fat (g)</th>
                    <th>Notes</th>
                </tr>
                {''.join([
                    f"<tr>"
                    f"<td>{meal.date.strftime('%Y-%m-%d')}</td>"
                    f"<td>{meal.name}</td>"
                    f"<td>{meal.calories}</td>"
                    f"<td>{meal.protein}</td>"
                    f"<td>{meal.carbs}</td>"
                    f"<td>{meal.fat}</td>"
                    f"<td>{meal.notes}</td>"
                    f"</tr>"
                    for meal in meals
                ])}
            </table>
        </div>
        
        <div class="section">
            <h2>Workout Logs</h2>
            <table>
                <tr>
                    <th>Date</th>
                    <th>Type</th>
                    <th>Duration (min)</th>
                    <th>Calories Burned</th>
                    <th>Notes</th>
                </tr>
                {''.join([
                    f"<tr>"
                    f"<td>{workout.date.strftime('%Y-%m-%d')}</td>"
                    f"<td>{workout.type}</td>"
                    f"<td>{workout.duration}</td>"
                    f"<td>{workout.calories_burned}</td>"
                    f"<td>{workout.notes}</td>"
                    f"</tr>"
                    for workout in workouts
                ])}
            </table>
        </div>
        
        <div class="section">
            <h2>Weight Logs</h2>
            <table>
                <tr>
                    <th>Date</th>
                    <th>Weight (kg)</th>
                    <th>Notes</th>
                </tr>
                {''.join([
                    f"<tr>"
                    f"<td>{weight.date.strftime('%Y-%m-%d')}</td>"
                    f"<td>{weight.weight}</td>"
                    f"<td>{weight.notes}</td>"
                    f"</tr>"
                    for weight in weights
                ])}
            </table>
        </div>
    </body>
    </html>
    """
    
    # Create PDF
    pdf = BytesIO()
    pisa.CreatePDF(html, dest=pdf)
    pdf.seek(0)
    return pdf

def generate_excel_report(user, meals, workouts, weights):
    # Create a new workbook
    wb = Workbook()
    
    # User Profile Sheet
    ws_profile = wb.active
    ws_profile.title = "Profile"
    ws_profile.append(["User Profile"])
    ws_profile.append(["Name", user.name])
    ws_profile.append(["Age", user.age])
    ws_profile.append(["Gender", user.gender])
    ws_profile.append(["Height (cm)", user.height])
    ws_profile.append(["Weight (kg)", user.weight])
    ws_profile.append(["Goal Weight (kg)", user.goal_weight])
    ws_profile.append(["Daily Calorie Target", user.daily_calories])
    
    # Meals Sheet
    ws_meals = wb.create_sheet("Meals")
    ws_meals.append(["Date", "Meal", "Calories", "Protein (g)", "Carbs (g)", "Fat (g)", "Notes"])
    for meal in meals:
        ws_meals.append([
            meal.date.strftime('%Y-%m-%d'),
            meal.name,
            meal.calories,
            meal.protein,
            meal.carbs,
            meal.fat,
            meal.notes
        ])
    
    # Workouts Sheet
    ws_workouts = wb.create_sheet("Workouts")
    ws_workouts.append(["Date", "Type", "Duration (min)", "Calories Burned", "Notes"])
    for workout in workouts:
        ws_workouts.append([
            workout.date.strftime('%Y-%m-%d'),
            workout.type,
            workout.duration,
            workout.calories_burned,
            workout.notes
        ])
    
    # Weights Sheet
    ws_weights = wb.create_sheet("Weights")
    ws_weights.append(["Date", "Weight (kg)", "Notes"])
    for weight in weights:
        ws_weights.append([
            weight.date.strftime('%Y-%m-%d'),
            weight.weight,
            weight.notes
        ])
    
    # Save to BytesIO
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_data.seek(0)
    return excel_data